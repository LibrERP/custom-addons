# © 2015-2023 Andrei Levin - Didotech srl (www.didotech.com)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html)

from odoo import _
from odoo import exceptions
from odoo.api import Environment
from dataclasses import dataclass
from collections import OrderedDict
from openpyxl.utils.cell import column_index_from_string


def column(letter):
    """Convert a column name into a numerical index
    ('A' -> 0)
    """
    return column_index_from_string(letter) - 1


class TableRow:
    _mapping: OrderedDict = {}
    env: Environment = None

    def __init__(self, env, data: list ):
        """
        Example:
        _mapping = OrderedDict([
            ('default_code', Column.B),
            ('serial_number', Column.E),
            ('imei', Column.F)
        ])
        """
        for key, column in self._mapping.items():
            value = data[column]
            if isinstance(value, str):
                value = value.strip()
            setattr(self, key, value)

        self.env = env

    def json(self):
        values = self.__dict__.copy()
        del values['env']
        return values


def import_sheet_generator(filename, content, header_lines_count=1, delimiter=','):
    name, file_type = filename.rsplit('.', 1)

    if file_type in ('xls', 'xlsb'):
        # "Excel"
        import xlrd

        for encoding in ('utf-8', 'latin-1', 'cp1252'):
            try:
                book = xlrd.open_workbook(file_contents=content, encoding_override=encoding)
                break
            except UnicodeDecodeError:
                pass
        else:
            raise exceptions.Warning(_('Error: Unknown encoding'))

        table = []
        sh = book.sheet_by_index(0)

        for rx in range(sh.nrows):
            row = []
            for cx in range(sh.ncols):
                row.append(sh.cell(rowx=rx, colx=cx).value)
            yield row
    elif file_type in ('xlsx', 'xlsm', 'xltx', 'xltm'):
        import openpyxl

        from io import BytesIO
        ## Create virtual File:
        virtual_file = BytesIO(content)
        book = openpyxl.load_workbook(virtual_file, read_only=True)
        # book = openpyxl.load_workbook(content, read_only=True)

        sh = book.worksheets[0]
        max_column = sh.max_column

        if max_column:
            for shrow in sh.rows:
                # row = []
                lrrow = list(shrow)
                for cx in range(max_column-1, 0, -1):
                    if lrrow[cx].value:
                        max_column = cx + 1
                        # Evaluates max columns to use in range
                        break
                break
        else:
            for y, row in enumerate(sh.rows, start=1):
                if y == header_lines_count:
                    break

            max_column = len(row) - 1

        for count, shrow in enumerate(sh.rows, start=1):
            row = []
            lrrow = list(shrow)

            # if len(shrow) - 1 < max_column:
            #     last_column = len(shrow) - 1
            # else:
            last_column = max_column

            for cx in range(0, last_column):
                row.append(lrrow[cx].value)

            if all(c is None for c in row) and count > header_lines_count:
                break
            yield row
    # elif file_type in ('ods', ):
    #     # "OpenOffice"
    #     from odoo.addons.core_extended.odf_to_array import ODSReader
    #     # import StringIO
    #     from io import BytesIO
    #
    #     ## Create virtual File:
    #     # virtual_file = StringIO.StringIO(content)
    #     virtual_file = BytesIO.StringIO(content)
    #
    #     book = ODSReader(virtual_file)
    #     table = book.sheet_by_index(0)
    #     for row in table:
    #         yield row
    # elif file_type == 'csv':
    #     # "CSV"
    #     import csv
    #
    #     def unicode_csv_reader(unicode_csv_data, dialect=csv.excel, **kwargs):
    #         # csv.py doesn't do Unicode; encode temporarily as UTF-8:
    #         # csv_reader = csv.reader(utf_8_encoder(unicode_csv_data),
    #         csv_reader = csv.reader(
    #             table_reader(unicode_csv_data),
    #             delimiter=delimiter,
    #             dialect=dialect, **kwargs)
    #         for row in csv_reader:
    #             # decode UTF-8 back to Unicode, cell by cell:
    #             # yield [str(cell, 'utf-8') for cell in row]
    #             yield row
    #
    #     # def utf_8_encoder(unicode_csv_data):
    #     #     for line in unicode_csv_data:
    #     #         yield line.encode('utf-8')
    #
    #     def table_reader(virtual_file_utf8):
    #         for line in virtual_file_utf8:
    #             yield line.decode('utf-8')
    #
    #     # from io import StringIO
    #     from io import BytesIO
    #     ## Create virtual File:
    #     # virtual_file = StringIO(content.decode('utf-8'))
    #     # virtual_file_utf8 = StringIO(content.decode('utf-8'))
    #
    #     virtual_file = BytesIO(content)
    #     virtual_file_utf8 = BytesIO(content)
    #
    #     ## Process CSV file:
    #     sample = virtual_file.read(512)
    #     virtual_file.seek(0)
    #     dialect = csv.Sniffer().sniff(sample.decode("utf-8"))
    #
    #     # table_latin1 = csv.reader(virtual_file, dialect)
    #     # self.table is an object of type '_csv.reader' and has no len() method
    #     # number_of_lines = sum(1 for row in table_latin1)
    #
    #     table = unicode_csv_reader(virtual_file_utf8, dialect)
    #     # table = table_reader(virtual_file_utf8, dialect)
    #
    #     # virtual_file.seek(0)
    #
    #     for row in unicode_csv_reader(virtual_file_utf8, dialect):
    #         yield row
    else:
        raise exceptions.Warning(_('Error: Unknown file extension'))
