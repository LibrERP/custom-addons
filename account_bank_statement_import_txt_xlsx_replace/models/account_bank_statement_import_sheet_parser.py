# Copyright 2019 ForgeFlow, S.L.
# Copyright 2020 CorporateHub (https://corporatehub.eu)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

import itertools
import logging
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from os import path

from odoo import _, api, exceptions, models

_logger = logging.getLogger(__name__)


class AccountBankStatementImportSheetParser(models.TransientModel):
    _name = 'account.bank.statement.import.sheet.parser'
    _description = 'Account Bank Statement Import Sheet Parser'

    @api.model
    def parse_header(self, data_file, filename, encoding, csv_options):
        """
        Parse table header

        :return: list of column names
        """
        # Parser = self.env['account.bank.statement.import.sheet.parser']
        csv_or_xlsx = self.read_table(data_file, filename, encoding, csv_options)
        return [str(value) for value in next(csv_or_xlsx)]

    # @api.model
    def parse_data_header(self, mapping, data_file, filename):
        header = {'name': '', 'date': False}
        finish_at = mapping.start_from and mapping.start_from - 1 or None
        header_data = list(
            self.read_table(
                data_file,
                filename,
                mapping.file_encoding,
                self.get_csv_options(mapping),
                start_from=0,
                finish_at=finish_at,
            )
        )

        if mapping.header_name_raw and mapping.header_name_column:
            header['name'] = header_data[int(mapping.header_name_raw)][
                int(mapping.header_name_column)
            ]

        if mapping.header_date_raw and mapping.header_date_column:
            header['date'] = header_data[int(mapping.header_date_raw)][
                int(mapping.header_date_column)
            ]

            if isinstance(header['date'], str):
                header['date'] = datetime.strptime(
                    header['date'], mapping.timestamp_format
                )

        return header

    @api.model
    def parse(self, mapping, data_file, filename):
        journal = self.env['account.journal'].browse(self.env.context.get('journal_id'))
        currency_code = (journal.currency_id or journal.company_id.currency_id).name
        account_number = journal.bank_account_id.acc_number

        if mapping.start_from:
            header = self.parse_data_header(mapping, data_file, filename)
        else:
            header = {}

        name = header.get('name', None) or _('%s: %s') % (
            journal.code,
            path.basename(filename),
        )
        lines = self._parse_lines(mapping, data_file, filename, currency_code)
        if not lines:
            return (
                currency_code,
                account_number,
                [
                    {
                        'name': name,
                        'transactions': [],
                    }
                ],
            )

        lines = list(sorted(lines, key=lambda line: line['timestamp']))
        first_line = lines[0]
        last_line = lines[-1]
        data = {
            'name': name,
            'date': header.get('date', None)
            and header['date'].date()
            or first_line['timestamp'].date(),
        }

        if mapping.balance_column:
            balance_start = first_line['balance']
            balance_start -= first_line['amount']
            balance_end = last_line['balance']
            data.update(
                {
                    'balance_start': float(balance_start),
                    'balance_end_real': float(balance_end),
                }
            )

        transactions = list(
            itertools.chain.from_iterable(
                map(lambda line: self._convert_line_to_transactions(line), lines)
            )
        )
        data.update(
            {
                'transactions': transactions,
            }
        )

        return currency_code, account_number, [data]

    @staticmethod
    def import_xls_generator(data_file, encoding, start_from=0, finish_at=None):
        import xlrd
        from xlrd.xldate import xldate_as_datetime

        try:
            book = xlrd.open_workbook(
                file_contents=data_file,
                encoding_override=(encoding if encoding else None),
            )
        except UnicodeDecodeError:
            pass

        sh = book.sheet_by_index(0)

        start_from = start_from and start_from - 1 or 0

        for rx in range(start_from, finish_at or sh.nrows):
            row = []
            for cx in range(sh.ncols):
                cell = sh.cell(rowx=rx, colx=cx)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    cell_value = xldate_as_datetime(cell.value, book.datemode)
                else:
                    cell_value = cell.value

                row.append(cell_value)
            yield row

    @staticmethod
    def import_xlsx_generator(data_file, start_from=0, finish_at=None):
        from openpyxl import load_workbook

        # Create virtual File:
        virtual_file = BytesIO(data_file)
        book = load_workbook(virtual_file, read_only=False)

        sh = book.worksheets[0]
        max_column = sh.max_column

        for count, shrow in enumerate(sh.rows, 1):
            if count >= start_from:
                lrrow = list(shrow)
                for cx in range(max_column - 1, 0, -1):
                    if lrrow[cx].value:
                        max_column = cx + 1
                        # Evaluates max columns to use in range
                        break
                break

        for count, shrow in enumerate(sh.rows, 1):
            if count >= start_from:
                if finish_at and count > finish_at - 1:
                    break
                row = []
                lrrow = list(shrow)
                for cx in range(0, max_column):
                    row.append(lrrow[cx].value)

                if all(c is None for c in row):
                    if finish_at and count < finish_at:
                        pass
                    else:
                        break
                yield row

    @staticmethod
    def import_csv_generator(data_file, encoding, csv_options, start_from=0, finish_at=None):
        import csv

        def unicode_csv_reader(
            unicode_csv_data,
            delimiter,
            start_from,
            finish_at,
            dialect=csv.excel,
            encoding='utf-8',
            **kwargs
        ):
            # csv.py doesn't do Unicode; encode temporarily as UTF-8:
            # csv_reader = csv.reader(utf_8_encoder(unicode_csv_data),
            csv_reader = csv.reader(
                table_reader(unicode_csv_data, start_from, finish_at, encoding),
                delimiter=delimiter,
                dialect=dialect,
                **kwargs
            )
            for row in csv_reader:
                # decode UTF-8 back to Unicode, cell by cell:
                # yield [str(cell, 'utf-8') for cell in row]
                row = [cell.strip() for cell in row]
                yield row

        def table_reader(virtual_file_utf8, start_from, finish_at, encoding):
            for count, line in enumerate(virtual_file_utf8, 1):
                if count >= start_from:
                    if finish_at and count > finish_at:
                        break

                    yield line.decode(encoding)

        # Create virtual File
        virtual_file_utf8 = BytesIO(data_file)

        for row in unicode_csv_reader(
            virtual_file_utf8,
            csv_options.get('delimiter', None),
            start_from,
            finish_at,
            encoding=encoding
        ):
            yield row

    def import_sheet_generator(
        self, data_file, filename, encoding, csv_options, start_from=0, finish_at=None
    ):
        name, file_type = filename.lower().rsplit('.', 1)
        if file_type in ('xls', 'xlsb'):
            return self.import_xls_generator(
                data_file, encoding, start_from=start_from, finish_at=finish_at
            )

        elif file_type in ('xlsx', 'xlsm', 'xltx', 'xltm'):
            return self.import_xlsx_generator(
                data_file, start_from=start_from, finish_at=finish_at
            )

        elif file_type == 'csv':
            return self.import_csv_generator(
                data_file, encoding, csv_options, start_from=start_from, finish_at=finish_at
            )

        else:
            raise exceptions.Warning(_('Error: Unknown file extension'))

    def read_table(
        self, data_file, filename, encoding, csv_options, start_from, finish_at=None
    ):
        # table = list(self.import_sheet_generator(data_file, filename, encoding, csv_options))
        table = self.import_sheet_generator(
            data_file,
            filename,
            encoding,
            csv_options,
            start_from=start_from,
            finish_at=finish_at,
        )
        return table

    def get_csv_options(self, mapping):
        csv_options = {}
        csv_delimiter = mapping._get_column_delimiter_character()
        if csv_delimiter:
            csv_options['delimiter'] = csv_delimiter
        if mapping.quotechar:
            csv_options['quotechar'] = mapping.quotechar

        return csv_options

    def _parse_lines(self, mapping, data_file, filename, currency_code):  # noqa: C901
        def get_index_or_none(header, column_name):
            if column_name and column_name in header:
                return header.index(column_name)
            else:
                return None

        def get_key_or_none(values_dict, key):
            return key and values_dict[key] or None

        def get_timestamp(values_dict, key, mapping):
            normalized_timestamp = values_dict[key]
            if normalized_timestamp:
                if isinstance(
                    normalized_timestamp, str
                ) and normalized_timestamp.replace('0', ''):
                    return datetime.strptime(
                        normalized_timestamp, mapping.timestamp_format
                    )
                elif isinstance(normalized_timestamp, str):
                    # Set date to minimum value accepted by JavaScript
                    return datetime(1970, 1, 1, 0, 0, 0)
                else:
                    # timestamp is already in datetime format. Do nothing
                    return normalized_timestamp
            else:
                return None

        def get_original_carrency_and_ammount(
            values, currency_key, ammount_key, currency, amount, mapping
        ):
            original_currency = get_key_or_none(values, currency_key)
            original_amount = get_key_or_none(values, ammount_key)
            if not original_currency:
                original_currency = currency
                original_amount = amount
            elif original_currency == currency:
                original_amount = amount

            line['original_currency'] = original_currency

            if original_amount:
                original_amount = self._parse_decimal(
                    original_amount, mapping
                ).copy_sign(amount)
            else:
                original_amount = 0.0

            return original_currency, original_amount

        def get_amount(values, columns, mapping):
            debit_amount = (
                columns['debit']
                and values[columns['debit']]
                and -abs(self._parse_decimal(values[columns['debit']], mapping))
                or 0
            )

            amount = values[columns['amount']] or debit_amount or 0
            amount = self._parse_decimal(amount, mapping)

            debit_credit = get_key_or_none(values, columns['debit_credit'])
            if debit_credit:
                amount = amount.copy_abs()
                if debit_credit == mapping.debit_value:
                    amount = -amount
            return amount

        if mapping.start_from:
            start_from = mapping.start_from
        else:
            start_from = 0

        csv_or_xlsx = self.read_table(
            data_file,
            filename,
            mapping.file_encoding,
            self.get_csv_options(mapping),
            start_from,
        )

        header = [str(value) for value in next(csv_or_xlsx)]

        columns = {
            'timestamp': header.index(mapping.timestamp_column),
            'currency': get_index_or_none(header, mapping.currency_column),
            'amount': header.index(mapping.amount_column),
            'debit': get_index_or_none(header, mapping.debit_column),
            'balance': get_index_or_none(header, mapping.balance_column),
            'original_currency': get_index_or_none(
                header, mapping.original_currency_column
            ),
            'original_amount': get_index_or_none(
                header, mapping.original_amount_column
            ),
            'debit_credit': get_index_or_none(header, mapping.debit_credit_column),
            'transaction_id': get_index_or_none(header, mapping.transaction_id_column),
            'description': get_index_or_none(header, mapping.description_column),
            'notes': get_index_or_none(header, mapping.notes_column),
            'reference': get_index_or_none(header, mapping.reference_column),
            'partner_name': get_index_or_none(header, mapping.partner_name_column),
            'bank_name': get_index_or_none(header, mapping.bank_name_column),
            'bank_account': get_index_or_none(header, mapping.bank_account_column),
        }

        # if isinstance(csv_or_xlsx, tuple):
        #     rows = range(1, csv_or_xlsx[1].nrows)
        # else:
        rows = csv_or_xlsx

        lines = []
        for row in rows:
            if row and any(row):
                values = list(row)

                currency = (
                    values[columns['currency']]
                    if columns['currency'] is not None
                    else currency_code
                )

                if currency.upper() == currency_code.upper():
                    # timestamp is a required field
                    timestamp = get_timestamp(values, columns['timestamp'], mapping)

                    if timestamp:
                        line = {
                            'timestamp': timestamp,
                            'amount': get_amount(values, columns, mapping),
                            'currency': currency,
                        }

                        balance = get_key_or_none(values, columns['balance'])
                        if balance:
                            line['balance'] = self._parse_decimal(balance, mapping)

                        (
                            original_currency,
                            original_amount,
                        ) = get_original_carrency_and_ammount(
                            values,
                            columns['original_currency'],
                            columns['original_amount'],
                            currency,
                            line['amount'],
                            mapping,
                        )
                        line.update(
                            {
                                'original_currency': original_currency,
                                'original_amount': original_amount,
                            }
                        )

                        for column in (
                            'transaction_id',
                            'description',
                            'notes',
                            'reference',
                            'partner_name',
                            'bank_name',
                            'bank_account',
                        ):
                            result = get_key_or_none(values, columns[column])
                            if result:
                                line[column] = result

                        lines.append(line)
                else:
                    _logger.warning('Wrong currency')
            else:
                break
        return lines

    @api.model
    def _convert_line_to_transactions(self, line):
        """Hook for extension"""
        timestamp = line['timestamp']
        amount = line['amount']
        currency = line['currency']
        original_amount = line['original_amount']
        original_currency = line['original_currency']
        transaction_id = line.get('transaction_id')
        description = line.get('description')
        notes = line.get('notes')
        reference = line.get('reference')
        partner_name = line.get('partner_name')
        bank_name = line.get('bank_name')
        bank_account = line.get('bank_account')

        transaction = {
            'date': timestamp,
            'amount': str(amount),
        }
        if currency != original_currency:
            original_currency = self.env['res.currency'].search(
                [('name', '=', original_currency)],
                limit=1,
            )
            if original_currency:
                transaction.update(
                    {
                        'amount_currency': str(original_amount),
                        'currency_id': original_currency.id,
                    }
                )

        if transaction_id:
            transaction['unique_import_id'] = '{}-{}'.format(
                transaction_id,
                int(timestamp.timestamp()),
            )

        transaction['name'] = description or _('N/A')
        if reference:
            transaction['ref'] = reference

        note = ''
        if bank_name:
            note += _('Bank: %s; ') % (bank_name,)
        if bank_account:
            note += _('Account: %s; ') % (bank_account,)
        if transaction_id:
            note += _('Transaction ID: %s; ') % (transaction_id,)
        if note and notes:
            note = '{}\n{}'.format(
                note,
                note.strip(),
            )
        elif note:
            note = note.strip()
        if note:
            transaction['note'] = note

        if partner_name:
            transaction['partner_name'] = partner_name
        if bank_account:
            transaction['account_number'] = bank_account

        return [transaction]

    @api.model
    def _parse_decimal(self, value, mapping):
        if isinstance(value, Decimal):
            return value
        elif isinstance(value, (float, int)):
            return Decimal(value)
        elif value:
            thousands, decimal = mapping._get_float_separators()
            value = value.replace(thousands, '')
            value = value.replace(decimal, '.')
            return Decimal(value)
